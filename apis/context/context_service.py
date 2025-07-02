import os
import uuid
import tempfile
import logging
from datetime import datetime
import pytz
from flask import g
from apis.utils.config import get_azure_blob_client, ensure_container_exists, get_aliased_blob_url
from apis.utils.databaseService import DatabaseService
from apis.utils.fileService import FileService

# Configure logging
logger = logging.getLogger(__name__)

# Define container for context files
CONTEXT_CONTAINER = "context-files"

class ContextService:
    @staticmethod
    def create_context(user_id, content=None, files=None, context_name=None, description=None):
        """
        Create a new context file with content and/or files
        
        Args:
            user_id (str): ID of the user creating the context
            content (str, optional): Text content for the context
            files (list, optional): List of file_ids to process and include
            context_name (str, optional): Name for the context
            description (str, optional): Description of the context
            
        Returns:
            tuple: (context_info, error)
                context_info is a dict with context metadata
                error is None on success, otherwise contains error message
        """
        try:
            # Ensure container exists
            ensure_container_exists(CONTEXT_CONTAINER)
            
            # Generate unique ID for the context
            context_id = str(uuid.uuid4())
            
            # Create a default name if none provided
            if not context_name:
                context_name = f"Context-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
            
            # Initialize content
            context_content = ""
            
            # Add provided content if any
            if content:
                context_content += content + "\n\n"
            
            # Process files if any
            file_processing_errors = []
            if files and isinstance(files, list) and len(files) > 0:
                logger.info(f"Processing {len(files)} files for context {context_id}")
                for file_id in files:
                    logger.info(f"Extracting content from file {file_id}")
                    file_content, error = ContextService.extract_file_content(file_id, user_id)
                    if error:
                        logger.error(f"Error extracting content from file {file_id}: {error}")
                        file_processing_errors.append(f"File {file_id}: {error}")
                        continue
                    
                    if file_content and file_content.strip():
                        context_content += f"--- Content from file ID: {file_id} ---\n"
                        context_content += file_content + "\n\n"
                        logger.info(f"Successfully extracted {len(file_content)} characters from file {file_id}")
                    else:
                        logger.warning(f"No content extracted from file {file_id}")
                        file_processing_errors.append(f"File {file_id}: No content extracted")
            
            # If no content after processing, return error
            if not context_content.strip():
                error_msg = "No content provided or extracted from files"
                if file_processing_errors:
                    error_msg += f". File processing errors: {'; '.join(file_processing_errors)}"
                return None, error_msg
            
            # Create a blob name using the context_id
            blob_name = f"{context_id}.txt"
            
            # Upload content to blob storage
            blob_service_client = get_azure_blob_client()
            container_client = blob_service_client.get_container_client(CONTEXT_CONTAINER)
            blob_client = container_client.get_blob_client(blob_name)
            
            # Convert content to bytes and upload
            blob_client.upload_blob(context_content.encode('utf-8'), overwrite=True)
            
            # Store context info in database
            db_conn = None
            cursor = None
            try:
                db_conn = DatabaseService.get_connection()
                cursor = db_conn.cursor()
                
                insert_query = """
                INSERT INTO context_files (
                    id, 
                    user_id, 
                    name, 
                    description,
                    path, 
                    created_at, 
                    modified_at,
                    file_size
                ) VALUES (?, ?, ?, ?, ?, DATEADD(HOUR, 2, GETUTCDATE()), DATEADD(HOUR, 2, GETUTCDATE()), ?)
                """
                
                cursor.execute(insert_query, [
                    context_id,
                    user_id,
                    context_name,
                    description,
                    blob_name,
                    len(context_content.encode('utf-8'))  # File size in bytes
                ])
                
                db_conn.commit()
            finally:
                if cursor:
                    cursor.close()
                if db_conn:
                    db_conn.close()
            
            # Return context info
            context_info = {
                "context_id": context_id,
                "name": context_name,
                "description": description,
                "file_size": len(context_content.encode('utf-8')),
                "created_at": datetime.now(pytz.UTC).isoformat(),
                "files_processed": len(files) if files else 0,
                "file_processing_errors": file_processing_errors if file_processing_errors else None
            }
            
            logger.info(f"Context created: {context_id} by user {user_id} with {len(context_content)} characters")
            return context_info, None
            
        except Exception as e:
            logger.error(f"Error creating context: {str(e)}")
            return None, str(e)
    
    @staticmethod
    def extract_file_content(file_id, user_id):
        """
        Extract content from a file using various extraction methods
        
        Args:
            file_id (str): ID of the file to extract content from
            user_id (str): ID of the user requesting extraction
            
        Returns:
            tuple: (content, error)
                content is the extracted text
                error is None on success, otherwise contains error message
        """
        try:
            logger.info(f"Starting content extraction for file {file_id}")
            
            # Get file info using FileService
            file_info, error = FileService.get_file_url(file_id, user_id)
            if error:
                logger.error(f"Error retrieving file info for {file_id}: {error}")
                return None, f"Error retrieving file: {error}"
            
            # Get file URL and name
            file_url = file_info.get('file_url')
            file_name = file_info.get('file_name', '')
            content_type = file_info.get('content_type', '')
            
            logger.info(f"File info - Name: {file_name}, Type: {content_type}, URL: {file_url}")
            
            # Create temporary file
            temp_file = None
            
            try:
                # Download the file
                import requests
                logger.info(f"Downloading file from URL: {file_url}")
                response = requests.get(file_url, timeout=30)
                response.raise_for_status()
                
                # Save to temporary file
                import tempfile
                temp_fd, temp_path = tempfile.mkstemp(suffix=f".{file_name.split('.')[-1]}" if '.' in file_name else '')
                temp_file = temp_path
                
                with os.fdopen(temp_fd, 'wb') as f:
                    f.write(response.content)
                
                logger.info(f"File downloaded to temporary path: {temp_path}")
                
                # Extract content based on file type
                extracted_content = None
                
                if file_name.lower().endswith('.txt') or content_type == 'text/plain':
                    logger.info("Processing as text file")
                    # Simple text file
                    with open(temp_path, 'r', encoding='utf-8', errors='ignore') as f:
                        extracted_content = f.read()
                        
                elif file_name.lower().endswith('.pdf'):
                    logger.info("Processing as PDF file")
                    extracted_content = ContextService._extract_pdf_content(temp_path)
                    
                elif file_name.lower().endswith(('.docx', '.doc')):
                    logger.info("Processing as Word document")
                    extracted_content = ContextService._extract_docx_content(temp_path)
                    
                elif file_name.lower().endswith('.csv'):
                    logger.info("Processing as CSV file")
                    extracted_content = ContextService._extract_csv_content(temp_path)
                    
                elif file_name.lower().endswith(('.xlsx', '.xls')):
                    logger.info("Processing as Excel file")
                    extracted_content = ContextService._extract_excel_content(temp_path)
                    
                elif file_name.lower().endswith('.py'):
                    logger.info("Processing as Python file")
                    extracted_content = ContextService._extract_python_content(temp_path)
                    
                elif file_name.lower().endswith('.ipynb'):
                    logger.info("Processing as Jupyter notebook")
                    extracted_content = ContextService._extract_notebook_content(temp_path)
                    
                else:
                    logger.warning(f"Unsupported file type: {content_type} for file {file_name}")
                    return None, f"Unsupported file type: {content_type}. Supported types: .txt, .pdf, .docx, .doc, .csv, .xlsx, .xls, .py, .ipynb"
                
                if extracted_content and extracted_content.strip():
                    logger.info(f"Successfully extracted {len(extracted_content)} characters from {file_name}")
                    return extracted_content, None
                else:
                    logger.warning(f"No content extracted from {file_name}")
                    return None, f"No content could be extracted from {file_name}"
                    
            finally:
                # Clean up temp file
                if temp_file and os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                        logger.info(f"Cleaned up temporary file: {temp_file}")
                    except Exception as cleanup_error:
                        logger.warning(f"Could not clean up temporary file {temp_file}: {cleanup_error}")
                    
        except Exception as e:
            logger.error(f"Error extracting file content from {file_id}: {str(e)}")
            return None, str(e)
    
    @staticmethod
    def _extract_pdf_content(file_path):
        """Extract text content from PDF file"""
        try:
            import PyPDF2
            
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text_content = ""
                
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    text_content += page.extract_text() + "\n"
                
                return text_content.strip()
                
        except ImportError:
            logger.error("PyPDF2 not installed. Install with: pip install PyPDF2")
            return None
        except Exception as e:
            logger.error(f"Error extracting PDF content: {str(e)}")
            # Try alternative method with fitz if available
            try:
                import fitz  # PyMuPDF
                doc = fitz.open(file_path)
                text_content = ""
                
                for page_num in range(doc.page_count):
                    page = doc[page_num]
                    text_content += page.get_text() + "\n"
                
                doc.close()
                return text_content.strip()
                
            except ImportError:
                logger.error("PyMuPDF not available. Install with: pip install PyMuPDF")
                return None
            except Exception as e2:
                logger.error(f"Error with PyMuPDF: {str(e2)}")
                return None
    
    @staticmethod
    def _extract_docx_content(file_path):
        """Extract text content from Word document"""
        try:
            from docx import Document
            
            doc = Document(file_path)
            text_content = ""
            
            for paragraph in doc.paragraphs:
                text_content += paragraph.text + "\n"
            
            return text_content.strip()
            
        except ImportError:
            logger.error("python-docx not installed. Install with: pip install python-docx")
            return None
        except Exception as e:
            logger.error(f"Error extracting DOCX content: {str(e)}")
            return None
    
    @staticmethod
    def _extract_csv_content(file_path):
        """Extract ALL content from CSV file"""
        try:
            import csv
            
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                # Try to detect delimiter
                sample = file.read(1024)
                file.seek(0)
                sniffer = csv.Sniffer()
                try:
                    delimiter = sniffer.sniff(sample).delimiter
                except:
                    delimiter = ','  # Default to comma
                
                reader = csv.reader(file, delimiter=delimiter)
                rows = list(reader)
                
                if not rows:
                    return "Empty CSV file"
                
                # Convert ALL data to readable format
                text_content = f"CSV Data with {len(rows)} rows:\n\n"
                
                # Add all rows
                for i, row in enumerate(rows):
                    # Clean and format row data
                    cleaned_row = [str(cell).strip() if cell is not None else "" for cell in row]
                    if i == 0:
                        text_content += f"Headers: {', '.join(cleaned_row)}\n"
                    else:
                        text_content += f"Row {i}: {', '.join(cleaned_row)}\n"
                
                return text_content
                
        except Exception as e:
            logger.error(f"Error extracting CSV content: {str(e)}")
            return None
    
    @staticmethod
    def _extract_excel_content(file_path):
        """Extract ALL content from Excel file"""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_content = f"Excel Workbook with {len(workbook.sheetnames)} sheets:\n\n"
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content += f"=== Sheet: {sheet_name} ===\n"
                
                # Get ALL data from sheet
                rows_with_data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None and str(cell).strip() for cell in row):
                        rows_with_data.append(row)
                
                if rows_with_data:
                    text_content += f"Total rows with data: {len(rows_with_data)}\n\n"
                    
                    # Add ALL rows
                    for i, row in enumerate(rows_with_data):
                        cleaned_row = [str(cell).strip() if cell is not None else "" for cell in row]
                        if i == 0:
                            text_content += f"Headers: {', '.join(cleaned_row)}\n"
                        else:
                            text_content += f"Row {i}: {', '.join(cleaned_row)}\n"
                else:
                    text_content += "(Empty sheet)\n"
                
                text_content += "\n"
            
            return text_content
            
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return None
        except Exception as e:
            logger.error(f"Error extracting Excel content: {str(e)}")
            return None
    
    @staticmethod
    def _extract_python_content(file_path):
        """Extract content from Python file"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                content = file.read()
                
                # Add header to identify it as Python code
                text_content = "=== Python Source Code ===\n\n"
                text_content += content
                
                return text_content
                
        except Exception as e:
            logger.error(f"Error extracting Python content: {str(e)}")
            return None
    
    @staticmethod
    def _extract_notebook_content(file_path):
        """Extract content from Jupyter notebook file"""
        try:
            import json
            
            with open(file_path, 'r', encoding='utf-8') as file:
                notebook = json.load(file)
            
            text_content = "=== Jupyter Notebook Content ===\n\n"
            
            if 'cells' in notebook:
                for i, cell in enumerate(notebook['cells']):
                    cell_type = cell.get('cell_type', 'unknown')
                    source = cell.get('source', [])
                    
                    # Convert source to string if it's a list
                    if isinstance(source, list):
                        source_text = ''.join(source)
                    else:
                        source_text = str(source)
                    
                    if source_text.strip():
                        text_content += f"--- Cell {i+1} ({cell_type}) ---\n"
                        text_content += source_text
                        text_content += "\n\n"
                        
                        # If it's a code cell, also include outputs if available
                        if cell_type == 'code' and 'outputs' in cell:
                            outputs = cell['outputs']
                            if outputs:
                                text_content += "Output:\n"
                                for output in outputs:
                                    if 'text' in output:
                                        output_text = output['text']
                                        if isinstance(output_text, list):
                                            output_text = ''.join(output_text)
                                        text_content += str(output_text) + "\n"
                                    elif 'data' in output and 'text/plain' in output['data']:
                                        output_text = output['data']['text/plain']
                                        if isinstance(output_text, list):
                                            output_text = ''.join(output_text)
                                        text_content += str(output_text) + "\n"
                                text_content += "\n"
            else:
                text_content += "No cells found in notebook\n"
            
            return text_content
            
        except ImportError:
            logger.error("json module not available")
            return None
        except json.JSONDecodeError as e:
            logger.error(f"Error parsing JSON notebook: {str(e)}")
            return None
        except Exception as e:
            logger.error(f"Error extracting notebook content: {str(e)}")
            return None
    
    @staticmethod
    def get_context(context_id, user_id=None, metadata_only=False):
        """
        Get a context file
        
        Args:
            context_id (str): ID of the context to retrieve
            user_id (str, optional): ID of the user requesting the context
            metadata_only (bool): Whether to return only metadata
            
        Returns:
            tuple: (context_data, error)
                context_data is a dict with context metadata and content
                error is None on success, otherwise contains error message
        """
        try:
            # First get metadata from database
            conn = None
            cursor = None
            try:
                conn = DatabaseService.get_connection()
                cursor = conn.cursor()
                
                query = """
                SELECT 
                    c.id, 
                    c.user_id, 
                    c.name, 
                    c.description, 
                    c.path, 
                    c.created_at, 
                    c.modified_at, 
                    c.file_size,
                    u.user_name
                FROM 
                    context_files c
                LEFT JOIN 
                    users u ON c.user_id = u.id
                WHERE 
                    c.id = ?
                """
                
                cursor.execute(query, [context_id])
                result = cursor.fetchone()
                
                if not result:
                    return None, f"Context with ID {context_id} not found"
                
                # Check user permissions if user_id provided
                context_user_id = result[1]
                if user_id and context_user_id != user_id:
                    # Check if requesting user is admin
                    cursor.execute("SELECT scope FROM users WHERE id = ?", [user_id])
                    user_scope = cursor.fetchone()
                    
                    # If not admin, deny access
                    if not user_scope or user_scope[0] != 0:
                        return None, "You don't have permission to access this context"
                
                # Build context metadata
                context_data = {
                    "context_id": result[0],
                    "owner_id": result[1],
                    "owner_name": result[8] if result[8] else "Unknown",
                    "name": result[2],
                    "description": result[3],
                    "path": result[4],
                    "created_at": result[5].isoformat() if result[5] else None,
                    "modified_at": result[6].isoformat() if result[6] else None,
                    "file_size": result[7]
                }
                
                # If only metadata requested, return now
                if metadata_only:
                    return context_data, None
                
                # Get context content from blob storage
                blob_name = result[4]
                blob_service_client = get_azure_blob_client()
                container_client = blob_service_client.get_container_client(CONTEXT_CONTAINER)
                blob_client = container_client.get_blob_client(blob_name)
                
                # Download content
                content_bytes = blob_client.download_blob().readall()
                content = content_bytes.decode('utf-8')
                
                # Add content to response
                context_data["content"] = content
                
                return context_data, None
                
            finally:
                if cursor:
                    cursor.close()
                if conn:
                    conn.close()
                    
        except Exception as e:
            logger.error(f"Error retrieving context: {str(e)}")
            return None, str(e)
    
    @staticmethod
    def update_context(context_id, user_id, content=None, files=None, append=True, context_name=None, description=None):
        """
        Update an existing context file
        
        Args:
            context_id (str): ID of the context to update
            user_id (str): ID of the user updating the context
            content (str, optional): New text content
            files (list, optional): List of file_ids to process and include
            append (bool): Whether to append to existing content (True) or replace (False)
            context_name (str, optional): New name for the context
            description (str, optional): New description for the context
            
        Returns:
            tuple: (updated_context, error)
                updated_context is a dict with updated context metadata
                error is None on success, otherwise contains error message
        """
        try:
            # First get existing context and check permissions
            existing_context, error = ContextService.get_context(context_id, user_id)
            if error:
                return None, error
            
            # Initialize updated content
            if append:
                # Get existing content
                updated_content = existing_context["content"]
                
                # Add separator if content or files being added
                if content or (files and len(files) > 0):
                    updated_content += "\n\n--- Updated on " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " ---\n\n"
            else:
                # Replace content
                updated_content = ""
            
            # Add new content if provided
            if content:
                updated_content += content + "\n\n"
            
            # Process files if any
            if files and isinstance(files, list) and len(files) > 0:
                for file_id in files:
                    file_content, error = ContextService.extract_file_content(file_id, user_id)
                    if error:
                        logger.warning(f"Error extracting content from file {file_id}: {error}")
                        continue
                    
                    if file_content:
                        updated_content += f"--- Content from file ID: {file_id} ---\n"
                        updated_content += file_content + "\n\n"
            
            # If no content update was requested and no files processed, just update metadata
            if not content and not files and (context_name or description is not None):
                metadata_only = True
            else:
                metadata_only = False
                
                # If no content after processing, return error
                if not updated_content.strip():
                    return None, "No content provided or extracted from files"
            
            # Update blob if content changed
            if not metadata_only:
                blob_name = existing_context["path"]
                blob_service_client = get_azure_blob_client()
                container_client = blob_service_client.get_container_client(CONTEXT_CONTAINER)
                blob_client = container_client.get_blob_client(blob_name)
                
                # Convert content to bytes and upload
                blob_client.upload_blob(updated_content.encode('utf-8'), overwrite=True)
            
            # Update database record
            db_conn = None
            cursor = None
            try:
                db_conn = DatabaseService.get_connection()
                cursor = db_conn.cursor()
                
                # Build update query based on provided fields
                update_fields = []
                params = []
                
                if context_name:
                    update_fields.append("name = ?")
                    params.append(context_name)
                
                if description is not None:  # Allow empty description
                    update_fields.append("description = ?")
                    params.append(description)
                
                if not metadata_only:
                    update_fields.append("file_size = ?")
                    params.append(len(updated_content.encode('utf-8')))
                
                # Always update modified_at
                update_fields.append("modified_at = DATEADD(HOUR, 2, GETUTCDATE())")
                
                # Only update if there are fields to update
                if update_fields:
                    update_query = f"""
                    UPDATE context_files
                    SET {', '.join(update_fields)}
                    WHERE id = ?
                    """
                    
                    # Add context_id to params
                    params.append(context_id)
                    
                    cursor.execute(update_query, params)
                    db_conn.commit()
                
                # Get updated context for response
                cursor.execute("""
                SELECT 
                    id, name, description, path, created_at, modified_at, file_size
                FROM 
                    context_files
                WHERE 
                    id = ?
                """, [context_id])
                
                updated = cursor.fetchone()
                
                updated_context = {
                    "context_id": updated[0],
                    "name": updated[1],
                    "description": updated[2],
                    "file_size": updated[6],
                    "created_at": updated[4].isoformat() if updated[4] else None,
                    "modified_at": updated[5].isoformat() if updated[5] else None,
                    "updated_content": not metadata_only,
                    "updated_metadata": bool(context_name or description is not None)
                }
                
                logger.info(f"Context updated: {context_id} by user {user_id}")
                return updated_context, None
                
            finally:
                if cursor:
                    cursor.close()
                if db_conn:
                    db_conn.close()
            
        except Exception as e:
            logger.error(f"Error updating context: {str(e)}")
            return None, str(e)
    
    @staticmethod
    def delete_context(context_id, user_id):
        """
        Delete a context file
        
        Args:
            context_id (str): ID of the context to delete
            user_id (str): ID of the user deleting the context
            
        Returns:
            tuple: (success, error)
                success is a boolean indicating if deletion was successful
                error is None on success, otherwise contains error message
        """
        try:
            # First check if context exists and user has permission
            existing_context, error = ContextService.get_context(context_id, user_id, metadata_only=True)
            if error:
                return False, error
            
            # Delete from blob storage
            blob_name = existing_context["path"]
            blob_service_client = get_azure_blob_client()
            container_client = blob_service_client.get_container_client(CONTEXT_CONTAINER)
            blob_client = container_client.get_blob_client(blob_name)
            
            # Delete the blob
            blob_client.delete_blob()
            
            # Delete database record
            db_conn = None
            cursor = None
            try:
                db_conn = DatabaseService.get_connection()
                cursor = db_conn.cursor()
                
                delete_query = """
                DELETE FROM context_files
                WHERE id = ?
                """
                
                cursor.execute(delete_query, [context_id])
                db_conn.commit()
                
                logger.info(f"Context deleted: {context_id} by user {user_id}")
                return True, None
                
            finally:
                if cursor:
                    cursor.close()
                if db_conn:
                    db_conn.close()
            
        except Exception as e:
            logger.error(f"Error deleting context: {str(e)}")
            return False, str(e)
    
    @staticmethod
    def list_contexts(user_id, include_shared=True, filter_name=None, limit=50, offset=0):
        """
        List all contexts available to a user
        
        Args:
            user_id (str): ID of the user listing contexts
            include_shared (bool): Whether to include shared contexts
            filter_name (str, optional): Filter contexts by name
            limit (int): Maximum number of contexts to return
            offset (int): Pagination offset
            
        Returns:
            tuple: (contexts, error)
                contexts is a list of context metadata
                error is None on success, otherwise contains error message
        """
        try:
            # Get user's scope (admin or not)
            conn = DatabaseService.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("SELECT scope FROM users WHERE id = ?", [user_id])
            user_scope_result = cursor.fetchone()
            
            if not user_scope_result:
                return None, "User not found"
                
            user_scope = user_scope_result[0]
            is_admin = (user_scope == 0)
            
            # Build query based on permissions
            if is_admin:
                # Admins can see all contexts
                query = """
                SELECT 
                    c.id, 
                    c.user_id, 
                    c.name, 
                    c.description, 
                    c.created_at, 
                    c.modified_at, 
                    c.file_size,
                    u.user_name
                FROM 
                    context_files c
                LEFT JOIN 
                    users u ON c.user_id = u.id
                """
                params = []
            else:
                # Regular users can only see their own contexts
                query = """
                SELECT 
                    c.id, 
                    c.user_id, 
                    c.name, 
                    c.description, 
                    c.created_at, 
                    c.modified_at, 
                    c.file_size,
                    u.user_name
                FROM 
                    context_files c
                LEFT JOIN 
                    users u ON c.user_id = u.id
                WHERE 
                    c.user_id = ?
                """
                params = [user_id]
            
            # Add name filter if provided
            if filter_name:
                query += " AND c.name LIKE ?"
                params.append(f"%{filter_name}%")
            
            # Add ordering
            query += " ORDER BY c.modified_at DESC"
            
            # Add pagination
            query += " OFFSET ? ROWS FETCH NEXT ? ROWS ONLY"
            params.extend([offset, limit])
            
            # Execute query
            cursor.execute(query, params)
            results = cursor.fetchall()
            
            # Format results
            contexts = []
            for result in results:
                contexts.append({
                    "context_id": result[0],
                    "owner_id": result[1],
                    "owner_name": result[7] if result[7] else "Unknown",
                    "name": result[2],
                    "description": result[3],
                    "created_at": result[4].isoformat() if result[4] else None,
                    "modified_at": result[5].isoformat() if result[5] else None,
                    "file_size": result[6]
                })
            
            cursor.close()
            conn.close()
            
            return contexts, None
            
        except Exception as e:
            logger.error(f"Error listing contexts: {str(e)}")
            return None, str(e)
